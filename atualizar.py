#!/usr/bin/env python3
"""
atualizar.py — Gerador automático do worker.js do Dashboard BI Qualitor
Uso:
    python atualizar.py \
        --projetos  Relatório_de_Projetos.xlsx \
        --contratos Renovação_de_Contratos.xlsx \
        --reducao   Redução_Contratual.xlsx \
        --template  template.js \
        --saida     worker.js
"""

import argparse
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("❌ openpyxl não instalado. Execute: pip install openpyxl")

# ── Helpers ──────────────────────────────────────────────────────────────────

def parse_dt(v):
    if v is None: return None
    if isinstance(v, (date, datetime)):
        return v.date() if isinstance(v, datetime) else v
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
        try: return datetime.strptime(str(v).strip(), fmt).date()
        except: pass
    return None

def pv(v):
    if v is None: return 0.0
    try: return float(v)
    except: return 0.0

def fmt(v):
    """Formata número como moeda BRL: R$ 1.234,56"""
    return "R$ {:,.2f}".format(v).replace(',', 'X').replace('.', ',').replace('X', '.')

def fmt_n(v):
    """Formata número sem prefixo R$: 1.234,56"""
    return "{:,.2f}".format(v).replace(',', 'X').replace('.', ',').replace('X', '.')

def aba_key(nome):
    m = re.match(r'S(\d+)M(\d+)', nome, re.IGNORECASE)
    return int(m.group(2)) * 10 + int(m.group(1)) if m else -1

def ultima_aba(wb):
    abas = sorted([a for a in wb.sheetnames if re.match(r'S\d+M\d+', a)], key=aba_key)
    if not abas:
        sys.exit("❌ Nenhuma aba SxMy encontrada no arquivo.")
    return abas[-2] if len(abas) >= 2 else None, abas[-1]

MESES = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun',
         'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']

EXCLUIR_ETAPA = {
    'Projeto Reprovado', 'Projeto Cancelado',
    'Cancelado', 'Desistência', 'Projeto Faturado'
}

BG = ["#ffffff", "#fafafa"]

def dash_cancel(v):
    if not v: return '<span style="color:#d1d5db;">—</span>'
    return f'<span style="color:#dc2626;font-weight:700;">−{fmt(v)}</span>'

def dash_red(v):
    if not v: return '<span style="color:#d1d5db;">—</span>'
    return f'<span style="color:#d97706;font-weight:700;">−{fmt(v)}</span>'

def dash_delta(v):
    if not v: return '<span style="color:#d1d5db;">—</span>'
    s = fmt_n(abs(v))
    if v > 0: return f'<span style="color:#059669;font-weight:700;">+R$ {s}</span>'
    return f'<span style="color:#dc2626;font-weight:700;">−R$ {s}</span>'

def brl(v):
    if not v: return '<span style="color:#d1d5db;">—</span>'
    return fmt(v)


# ── 1. PIPELINE ───────────────────────────────────────────────────────────────

def calcular_pipeline(path_proj):
    wb = openpyxl.load_workbook(path_proj, read_only=True, data_only=True)
    aba_ant, aba_nova = ultima_aba(wb)
    print(f"  Pipeline: {aba_ant} → {aba_nova}")

    def rows(aba): return [r for r in wb[aba].iter_rows(min_row=2, values_only=True) if r[0]]

    nova = rows(aba_nova)
    ant  = rows(aba_ant) if aba_ant else []
    hoje = date.today()

    # Filtros
    ativos   = [r for r in nova if str(r[6] or '') not in EXCLUIR_ETAPA and str(r[5] or '') != 'Cancelado']
    enviados = [r for r in nova if str(r[6] or '') == 'Projeto Enviado']
    neg      = [r for r in nova if str(r[6] or '') in ('Negociação', 'Em Negociação')]
    pf       = [r for r in nova if str(r[6] or '') == 'Processo Financeiro']
    fechados = sorted(
        [r for r in nova if str(r[6] or '') == 'Projeto Faturado' and parse_dt(r[9]) and parse_dt(r[9]).year >= 2026],
        key=lambda r: parse_dt(r[9]) or date.min, reverse=True
    )

    n_fat   = len(fechados)
    vt_fat  = sum(pv(r[12]) for r in fechados)
    mrr_fat = sum(pv(r[13]) for r in fechados)
    arr_fat = sum(pv(r[14]) for r in fechados)
    vt_env  = sum(pv(r[12]) for r in enviados)
    vt_neg  = sum(pv(r[12]) for r in neg)

    # VT Enviados formatado em M
    vt_env_fmt = f"R$ {vt_env/1_000_000:.2f}M".replace('.', ',')

    # Responsáveis com card
    RESP_CARDS = ['Pedro Schreck', 'Anderson Bamberg', 'Talita Rodrigues', 'Ines Cristine Nunes Diogo']
    total_ativos = len(ativos)
    resp_counts = {}
    resp_proj   = {r: [] for r in RESP_CARDS}

    for r in ativos:
        resp = str(r[3] or '').strip()
        resp_counts[resp] = resp_counts.get(resp, 0) + 1
        if resp in RESP_CARDS:
            dt_ab = parse_dt(r[1])
            dias  = (hoje - dt_ab).days if dt_ab else None
            resp_proj[resp].append({
                'np': str(r[0]),
                'cliente': str(r[2] or ''),
                'etapa': str(r[6] or ''),
                'vt': round(pv(r[12]), 2),
                'abertura': dt_ab.strftime('%d/%m/%Y') if dt_ab else '',
                'dias': dias
            })

    for k in resp_proj:
        resp_proj[k].sort(key=lambda x: (-x['vt'], -(x['dias'] or 0)))

    resp_data = {}
    for resp in RESP_CARDS:
        qtd = resp_counts.get(resp, 0)
        resp_data[resp] = {
            'qtd': qtd,
            'pct': round(qtd / total_ativos * 100, 1) if total_ativos else 0,
            'projetos': resp_proj[resp]
        }

    # Movimentações PF
    pf_ant_set = {str(r[0]) for r in ant if str(r[6] or '') == 'Processo Financeiro'}
    pf_nova_set = {str(r[0]) for r in nova}
    idx_nova = {str(r[0]): r for r in nova}
    idx_ant  = {str(r[0]): r for r in ant}

    ent_pf = [r for r in nova if str(r[6] or '') == 'Processo Financeiro' and str(r[0]) not in pf_ant_set]
    sai_pf = []
    for np, r_ant in {str(r[0]): r for r in ant if str(r[6] or '') == 'Processo Financeiro'}.items():
        r_at = idx_nova.get(np)
        if r_at and str(r_at[6] or '') != 'Processo Financeiro':
            sai_pf.append({'np': np, 'cli': str(r_ant[2] or ''), 'para': str(r_at[6] or ''), 'vt': pv(r_at[12])})

    # Fechados por mês
    fat_mes = {}
    for r in fechados:
        dt = parse_dt(r[9])
        mk = MESES[dt.month - 1] + '/26'
        fat_mes.setdefault(mk, {'n': 0, 'vt': 0})
        fat_mes[mk]['n'] += 1
        fat_mes[mk]['vt'] += pv(r[12])

    return {
        'aba_ant': aba_ant, 'aba_nova': aba_nova,
        'n_ativos': total_ativos,
        'n_enviados': len(enviados),
        'n_neg': len(neg), 'vt_neg': vt_neg,
        'vt_env_fmt': vt_env_fmt,
        'n_fat': n_fat, 'vt_fat': vt_fat, 'mrr_fat': mrr_fat, 'arr_fat': arr_fat,
        'resp_data': resp_data,
        'resp_counts': resp_counts,
        'fechados': fechados, 'fat_mes': fat_mes,
        'neg': neg, 'pf': pf,
        'ent_pf': ent_pf, 'sai_pf': sai_pf,
    }


# ── 2. GESTÃO ─────────────────────────────────────────────────────────────────

def calcular_gestao(path_contratos, path_reducao):
    wb_ren = openpyxl.load_workbook(path_contratos, read_only=True, data_only=True)
    _, aba_nova = ultima_aba(wb_ren)
    print(f"  Gestão: aba {aba_nova}")

    def rows(ws): return [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]

    base_rows = rows(wb_ren['Renovação de Contratos'])
    nova_rows = rows(wb_ren[aba_nova])
    hoje = date.today()

    # KPIs contagem
    cancelados_g = [r for r in nova_rows if str(r[6] or '') == 'Contrato Cancelado']
    carteira_at  = [r for r in nova_rows if str(r[5] or '') != 'Cancelado' and str(r[6] or '') != 'Contrato Cancelado']
    renovados    = [r for r in carteira_at if str(r[6] or '') == 'Renovado']
    pf_g         = [r for r in carteira_at if str(r[6] or '') == 'Processo Financeiro']
    em_analise   = [r for r in carteira_at if str(r[6] or '') == 'Em análise pelo cliente']
    analise_req  = [r for r in carteira_at if str(r[6] or '') == 'Análise de Requisição']

    n_jan    = len(base_rows)
    n_atual  = len(carteira_at)
    n_cancel = len(cancelados_g)
    n_renov  = len(renovados)
    vtc_jan  = sum(pv(r[9]) for r in base_rows)
    reducao  = sum(pv(r[9]) for r in cancelados_g)
    vtc_at   = vtc_jan - reducao
    ticket   = vtc_at / n_atual if n_atual else 0

    # MRR/ARR por mês (aba base = vencimento)
    mrr_mes = {m: {'n': 0, 'v': 0, 'cancel_mrr': 0, 'red_mrr': 0} for m in MESES}
    arr_mes = {m: {'n': 0, 'v': 0, 'cancel_arr': 0, 'red_arr': 0} for m in MESES}

    for r in base_rows:
        dt = parse_dt(r[7])
        if not dt or dt.year < 2026: continue
        m = MESES[dt.month - 1]
        mrr = pv(r[18]); arr = pv(r[19])
        if mrr > 0: mrr_mes[m]['n'] += 1; mrr_mes[m]['v'] += mrr
        if arr > 0: arr_mes[m]['n'] += 1; arr_mes[m]['v'] += arr

    # Cancelamentos por mês (aba SxMy = mais atualizado)
    for r in nova_rows:
        dt_c = parse_dt(r[20])
        if not dt_c or dt_c.year < 2026: continue
        m = MESES[dt_c.month - 1]
        mrr_mes[m]['cancel_mrr'] += pv(r[18])
        arr_mes[m]['cancel_arr'] += pv(r[19])

    # Reduções por mês (Redução_Contratual.xlsx)
    if path_reducao and Path(path_reducao).exists():
        wb_red = openpyxl.load_workbook(path_reducao, read_only=True, data_only=True)
        _, aba_red = ultima_aba(wb_red)
        print(f"  Redução: aba {aba_red}")
        for r in wb_red[aba_red].iter_rows(min_row=2, values_only=True):
            if not r[0]: continue
            dt = parse_dt(r[9])
            if not dt or dt.year < 2026: continue
            m = MESES[dt.month - 1]
            mrr_mes[m]['red_mrr'] += pv(r[11])
            arr_mes[m]['red_arr'] += pv(r[12])

    total_mrr = sum(d['v'] for d in mrr_mes.values())
    total_arr = sum(d['v'] for d in arr_mes.values())
    total_cancel_mrr = sum(d['cancel_mrr'] for d in mrr_mes.values())
    total_cancel_arr = sum(d['cancel_arr'] for d in arr_mes.values())
    total_red_mrr    = sum(d['red_mrr'] for d in mrr_mes.values())
    total_red_arr    = sum(d['red_arr'] for d in arr_mes.values())
    n_mrr = sum(d['n'] for d in mrr_mes.values())
    n_arr = sum(d['n'] for d in arr_mes.values())

    # Semáforo Em Análise
    def semaforo(dt):
        if not dt: return '?', '?', '#6b7280', '#f3f4f6'
        dias = (hoje - dt).days
        if dias > 60:   return '🔴', f'{dias} dias atrás',    '#dc2626', '#fee2e2'
        elif dias > 30: return '🟡', f'{dias} dias atrás',    '#92400e', '#fef9c3'
        else:           return '🟢', f'vence em {abs(dias)} dias', '#166534', '#dcfce7'

    sem_count = {'🔴': 0, '🟡': 0, '🟢': 0}
    for r in em_analise:
        e, _, _, _ = semaforo(parse_dt(r[7]))
        sem_count[e] = sem_count.get(e, 0) + 1

    # Análise Req por mês
    req_mes = {}
    for r in analise_req:
        dt = parse_dt(r[7])
        if not dt: continue
        mk = MESES[dt.month - 1] + '/26'
        req_mes.setdefault(mk, {'n': 0, 'vtc': 0, 'mrr': 0, 'arr': 0})
        req_mes[mk]['n']   += 1
        req_mes[mk]['vtc'] += pv(r[9])
        req_mes[mk]['mrr'] += pv(r[18])
        req_mes[mk]['arr'] += pv(r[19])

    return {
        'aba_nova': aba_nova,
        'n_jan': n_jan, 'n_atual': n_atual,
        'n_cancel': n_cancel, 'n_renov': n_renov,
        'pct_cancel': round(n_cancel / len(nova_rows) * 100, 1) if nova_rows else 0,
        'pct_renov': round(n_renov / n_atual * 100, 1) if n_atual else 0,
        'vtc_jan': vtc_jan, 'reducao': reducao, 'vtc_at': vtc_at, 'ticket': ticket,
        'mrr_mes': mrr_mes, 'arr_mes': arr_mes,
        'total_mrr': total_mrr, 'total_arr': total_arr,
        'total_cancel_mrr': total_cancel_mrr, 'total_cancel_arr': total_cancel_arr,
        'total_red_mrr': total_red_mrr, 'total_red_arr': total_red_arr,
        'n_mrr': n_mrr, 'n_arr': n_arr,
        'renovados': renovados, 'pf_g': pf_g,
        'em_analise': em_analise, 'analise_req': analise_req,
        'cancelados_g': cancelados_g,
        'sem_count': sem_count,
        'req_mes': req_mes,
        'semaforo_fn': semaforo,
    }


# ── 3. RESUMO ────────────────────────────────────────────────────────────────

def calcular_resumo(pip, gest):
    """MRR ajustado mês a mês."""
    mrr_baseline = gest['total_mrr']

    # MRR aprovados por mês (projetos fechados)
    mrr_aprov_mes = {m: 0 for m in MESES}
    for r in pip['fechados']:
        dt = parse_dt(r[9])
        if dt: mrr_aprov_mes[MESES[dt.month - 1]] += pv(r[13])

    evolucao = []
    mrr_ajust = mrr_baseline
    ultimo_mes_com_mov = None

    for m in MESES:
        aprov  = mrr_aprov_mes[m]
        cancel = gest['mrr_mes'][m]['cancel_mrr']
        red    = gest['mrr_mes'][m]['red_mrr']
        delta  = aprov - cancel - red
        mrr_ajust += delta
        if aprov or cancel or red:
            ultimo_mes_com_mov = m
        evolucao.append((m, aprov, cancel, red, mrr_ajust, delta))

    variacao = mrr_ajust - mrr_baseline
    var_pct  = round(variacao / mrr_baseline * 100, 2) if mrr_baseline else 0
    mes_ref  = ultimo_mes_com_mov or MESES[0]
    mes_ref_label = MESES.index(mes_ref) + 1
    mes_ref_fmt = f"{mes_ref}/2026"

    return {
        'mrr_baseline': mrr_baseline,
        'mrr_ajust': mrr_ajust,
        'variacao': variacao,
        'var_pct': var_pct,
        'mes_ref_fmt': mes_ref_fmt,
        'evolucao': evolucao,
    }


# ── 4. GERADOR DE HTML ───────────────────────────────────────────────────────

def gerar_tbody_neg(neg):
    rows = ""
    for i, r in enumerate(neg):
        bg = BG[i % 2]
        np = str(r[0]); cli = str(r[2] or ''); resp = str(r[3] or '').split()[0]
        vt = pv(r[12]); etapa = str(r[6] or ''); cotacao = str(r[7] or '') if r[7] else '—'
        rows += (
            f'<tr style="background:{bg};border-bottom:1px solid #fef3c7;">'
            f'<td style="padding:10px 16px;font-family:monospace;font-size:11px;font-weight:600;color:#f59e0b;">{np}</td>'
            f'<td style="padding:10px 16px;font-size:12px;font-weight:600;color:#111827;">{cli}</td>'
            f'<td style="padding:10px 16px;font-size:11px;color:#6b7280;">{resp}</td>'
            f'<td style="padding:10px 16px;font-size:11px;color:#374151;">{etapa}</td>'
            f'<td style="padding:10px 16px;text-align:right;font-size:12px;font-weight:700;color:#111827;">{fmt(vt)}</td>'
            f'<td style="padding:10px 16px;font-size:11px;color:#6b7280;">{cotacao}</td>'
            f'</tr>'
        )
    return rows


def gerar_tbody_fat_mes(fat_mes, n_fat, vt_fat):
    rows = ""
    idx = 0
    for mk in [f'{m}/26' for m in MESES]:
        d = fat_mes.get(mk)
        if not d: continue
        bg = BG[idx % 2]; idx += 1
        rows += (
            f'<tr style="background:{bg};border-bottom:1px solid #f3f4f6;">'
            f'<td style="padding:12px 20px;font-size:14px;font-weight:700;color:#111827;">{mk}</td>'
            f'<td style="padding:12px 20px;text-align:center;"><span style="background:#f0f2f5;color:#374151;font-size:13px;font-weight:700;padding:3px 12px;border-radius:99px;">{d["n"]}</span></td>'
            f'<td style="padding:12px 20px;text-align:right;font-size:14px;font-weight:700;color:#111827;">{fmt(d["vt"])}</td>'
            f'</tr>'
        )
    rows += (
        f'<tr style="background:#f0fdf4;border-top:2px solid #bbf7d0;">'
        f'<td style="padding:12px 20px;font-size:12px;font-weight:800;text-transform:uppercase;color:#059669;">Total</td>'
        f'<td style="padding:12px 20px;text-align:center;font-size:14px;font-weight:800;color:#059669;">{n_fat}</td>'
        f'<td style="padding:12px 20px;text-align:right;font-size:14px;font-weight:800;color:#111827;">{fmt(vt_fat)}</td>'
        f'</tr>'
    )
    return rows


def gerar_tbody_detalhe(fechados, n_fat, vt_fat, mrr_fat, arr_fat):
    rows = ""
    for i, r in enumerate(fechados):
        dt = parse_dt(r[9])
        dt_s = dt.strftime('%d/%m/%Y') if dt else '—'
        mrr = pv(r[13]); arr = pv(r[14])
        bg = BG[i % 2]
        resp = str(r[3] or '').split()[0]
        rows += (
            f'<tr style="background:{bg};border-bottom:1px solid #f3f4f6;">'
            f'<td style="padding:8px 14px;font-family:monospace;font-size:11px;font-weight:600;color:#7c3aed;">{r[0]}</td>'
            f'<td style="padding:8px 14px;font-size:12px;font-weight:600;color:#111827;">{str(r[2] or "")}</td>'
            f'<td style="padding:8px 14px;font-size:11px;color:#6b7280;">{resp}</td>'
            f'<td style="padding:8px 14px;text-align:center;font-size:11px;color:#374151;">{dt_s}</td>'
            f'<td style="padding:8px 14px;text-align:right;font-size:12px;font-weight:600;color:#7c3aed;">{brl(mrr)}</td>'
            f'<td style="padding:8px 14px;text-align:right;font-size:12px;font-weight:600;color:#0891b2;">{brl(arr)}</td>'
            f'</tr>'
        )
    rows += (
        f'<tr style="background:#f0fdf4;border-top:2px solid #bbf7d0;">'
        f'<td colspan="3" style="padding:12px 16px;font-size:11px;font-weight:700;text-transform:uppercase;color:#059669;">Total · {n_fat} projetos</td>'
        f'<td style="padding:12px 16px;text-align:right;font-size:13px;font-weight:800;color:#111827;">{fmt(vt_fat)}</td>'
        f'<td style="padding:12px 16px;text-align:right;font-size:13px;font-weight:800;color:#7c3aed;">{fmt(mrr_fat)}</td>'
        f'<td style="padding:12px 16px;text-align:right;font-size:13px;font-weight:800;color:#0891b2;">{fmt(arr_fat)}</td>'
        f'</tr>'
    )
    return rows


def gerar_tbody_ent_pf(ent_pf):
    if not ent_pf:
        return '<tr style="background:#f9fafb;"><td colspan="4" style="padding:16px 14px;text-align:center;font-size:12px;color:#9ca3af;">Nenhuma entrada no Processo Financeiro esta semana</td></tr>'
    rows = ""
    tot = 0
    for i, r in enumerate(ent_pf):
        vt = pv(r[12]); tot += vt
        resp = str(r[3] or '').split()[0]
        bg = BG[i % 2]
        rows += (
            f'<tr style="background:{bg};border-bottom:1px solid #f3f4f6;">'
            f'<td style="padding:8px 14px;font-family:monospace;font-size:11px;font-weight:600;color:#7c3aed;">{r[0]}</td>'
            f'<td style="padding:8px 14px;font-size:12px;font-weight:600;color:#111827;">{str(r[2] or "")}</td>'
            f'<td style="padding:8px 14px;font-size:11px;color:#6b7280;">{resp}</td>'
            f'<td style="padding:8px 14px;text-align:right;font-size:12px;font-weight:700;color:#7c3aed;">{fmt(vt)}</td>'
            f'</tr>'
        )
    n = len(ent_pf)
    s = 's' if n != 1 else ''
    rows += (
        f'<tr style="background:#f5f3ff;border-top:2px solid #e9d5ff;">'
        f'<td colspan="3" style="padding:9px 14px;font-size:11px;font-weight:700;text-transform:uppercase;color:#7c3aed;">Total · {n} projeto{s}</td>'
        f'<td style="padding:9px 14px;text-align:right;font-size:13px;font-weight:800;color:#7c3aed;">{fmt(tot)}</td>'
        f'</tr>'
    )
    return rows


def gerar_tbody_sai_pf(sai_pf, aba_ant, aba_nova):
    if not sai_pf:
        return f'<tr style="background:#f9fafb;"><td colspan="5" style="padding:16px 14px;text-align:center;font-size:12px;color:#9ca3af;">Nenhuma movimentação de Processo Financeiro nesta semana ({aba_ant} → {aba_nova})</td></tr>'
    rows = ""
    for i, s in enumerate(sai_pf):
        bg = BG[i % 2]
        rows += (
            f'<tr style="background:{bg};border-bottom:1px solid #f3f4f6;">'
            f'<td style="padding:8px 14px;font-family:monospace;font-size:11px;font-weight:600;color:#dc2626;">{s["np"]}</td>'
            f'<td style="padding:8px 14px;font-size:12px;font-weight:600;color:#111827;">{s["cli"]}</td>'
            f'<td style="padding:8px 14px;font-size:11px;color:#6b7280;">{s["para"]}</td>'
            f'<td style="padding:8px 14px;text-align:right;font-size:12px;font-weight:700;color:#111827;">{fmt(s["vt"])}</td>'
            f'</tr>'
        )
    return rows


def gerar_tbody_mrr(mrr_mes, total_mrr, total_cancel_mrr, total_red_mrr, n_mrr):
    rows = ""
    idx = 0
    for m in MESES:
        d = mrr_mes[m]
        if d['v'] == 0 and d['cancel_mrr'] == 0 and d['red_mrr'] == 0: continue
        bg = BG[idx % 2]; idx += 1
        n_s = str(d['n']) if d['n'] else '—'
        v_s = fmt(d['v']) if d['v'] else '<span style="color:#d1d5db;">—</span>'
        rows += (
            f'<tr style="background:{bg};border-bottom:1px solid #f3f4f6;">'
            f'<td style="padding:10px 16px;font-size:13px;font-weight:700;color:#111827;">{m}/26</td>'
            f'<td style="padding:10px 16px;text-align:center;font-size:13px;color:#6b7280;">{n_s}</td>'
            f'<td style="padding:10px 16px;text-align:right;font-size:13px;font-weight:700;color:#2563eb;">{v_s}</td>'
            f'<td style="padding:10px 16px;text-align:right;font-size:12px;">{dash_cancel(d["cancel_mrr"])}</td>'
            f'<td style="padding:10px 16px;text-align:right;font-size:12px;">{dash_red(d["red_mrr"])}</td>'
            f'</tr>'
        )
    tc = fmt_n(total_cancel_mrr); tr = fmt_n(total_red_mrr)
    rows += (
        f'<tr style="background:#eff6ff;border-top:2px solid #bfdbfe;">'
        f'<td style="padding:10px 16px;font-size:11px;font-weight:800;text-transform:uppercase;color:#2563eb;">Total</td>'
        f'<td style="padding:10px 16px;text-align:center;font-size:13px;font-weight:800;color:#2563eb;">{n_mrr}</td>'
        f'<td style="padding:10px 16px;text-align:right;font-size:13px;font-weight:800;color:#2563eb;">{fmt(total_mrr)}</td>'
        f'<td style="padding:10px 16px;text-align:right;font-size:12px;"><span style="color:#dc2626;font-weight:800;">−R$ {tc}</span></td>'
        f'<td style="padding:10px 16px;text-align:right;font-size:12px;"><span style="color:#d97706;font-weight:800;">−R$ {tr}</span></td>'
        f'</tr>'
    )
    return rows


def gerar_tbody_arr(arr_mes, total_arr, total_cancel_arr, total_red_arr, n_arr):
    rows = ""
    idx = 0
    for m in MESES:
        d = arr_mes[m]
        if d['v'] == 0 and d['cancel_arr'] == 0 and d['red_arr'] == 0: continue
        bg = BG[idx % 2]; idx += 1
        n_s = str(d['n']) if d['n'] else '—'
        v_s = fmt(d['v']) if d['v'] else '<span style="color:#d1d5db;">—</span>'
        rows += (
            f'<tr style="background:{bg};border-bottom:1px solid #f3f4f6;">'
            f'<td style="padding:10px 16px;font-size:13px;font-weight:700;color:#111827;">{m}/26</td>'
            f'<td style="padding:10px 16px;text-align:center;font-size:13px;color:#6b7280;">{n_s}</td>'
            f'<td style="padding:10px 16px;text-align:right;font-size:13px;font-weight:700;color:#059669;">{v_s}</td>'
            f'<td style="padding:10px 16px;text-align:right;font-size:12px;">{dash_cancel(d["cancel_arr"])}</td>'
            f'<td style="padding:10px 16px;text-align:right;font-size:12px;">{dash_red(d["red_arr"])}</td>'
            f'</tr>'
        )
    tc = fmt_n(total_cancel_arr); tr = fmt_n(total_red_arr)
    rows += (
        f'<tr style="background:#f0fdf4;border-top:2px solid #bbf7d0;">'
        f'<td style="padding:10px 16px;font-size:11px;font-weight:800;text-transform:uppercase;color:#059669;">Total</td>'
        f'<td style="padding:10px 16px;text-align:center;font-size:13px;font-weight:800;color:#059669;">{n_arr}</td>'
        f'<td style="padding:10px 16px;text-align:right;font-size:13px;font-weight:800;color:#059669;">{fmt(total_arr)}</td>'
        f'<td style="padding:10px 16px;text-align:right;font-size:12px;"><span style="color:#dc2626;font-weight:800;">−R$ {tc}</span></td>'
        f'<td style="padding:10px 16px;text-align:right;font-size:12px;"><span style="color:#d1d5db;">—</span></td>'
        f'</tr>'
    )
    return rows


def gerar_tbody_gestao_tabela(registros, cor_np, cor_total, label_total, extra_col_fn=None):
    """Genérico para Renovados, PF, Cancelados."""
    rows = ""; tot = {'vtc': 0, 'mrr': 0, 'arr': 0}
    for i, r in enumerate(registros):
        dt = parse_dt(r[7])
        mk = (MESES[dt.month - 1] + '/26') if dt else '—'
        vtc = pv(r[9]); mrr = pv(r[18]); arr = pv(r[19])
        tot['vtc'] += vtc; tot['mrr'] += mrr; tot['arr'] += arr
        bg = BG[i % 2]
        extra = extra_col_fn(r) if extra_col_fn else ''
        rows += (
            f'<tr style="background:{bg};border-bottom:1px solid #f3f4f6;">'
            f'{extra}'
            f'<td style="padding:9px 14px;font-family:monospace;text-align:left;font-size:11px;font-weight:600;color:{cor_np};">{str(r[0])}</td>'
            f'<td style="padding:9px 14px;text-align:left;font-size:12px;font-weight:600;color:#111827;">{str(r[2] or "")}</td>'
            f'<td style="padding:9px 14px;text-align:left;font-size:11px;font-weight:600;color:#6b7280;">{str(r[3] or "").split()[0]}</td>'
            f'<td style="padding:9px 14px;text-align:center;font-size:11px;font-weight:600;color:#059669;">{mk}</td>'
            f'<td style="padding:9px 14px;text-align:right;font-size:12px;font-weight:700;color:#111827;">{fmt(vtc)}</td>'
            f'<td style="padding:9px 14px;text-align:right;font-size:12px;font-weight:700;color:#7c3aed;">{brl(mrr)}</td>'
            f'<td style="padding:9px 14px;text-align:right;font-size:12px;font-weight:700;color:#0891b2;">{brl(arr)}</td>'
            f'</tr>'
        )
    n = len(registros)
    colspan = "4" if not extra_col_fn else "5"
    rows += (
        f'<tr style="background:{cor_total[0]};border-top:2px solid {cor_total[1]};">'
        f'<td colspan="{colspan}" style="padding:9px 14px;font-size:11px;font-weight:800;text-transform:uppercase;color:{cor_total[2]};">{label_total} · {n} contratos</td>'
        f'<td style="padding:9px 14px;text-align:right;font-size:13px;font-weight:800;color:#111827;">{fmt(tot["vtc"])}</td>'
        f'<td style="padding:9px 14px;text-align:right;font-size:13px;font-weight:800;color:#7c3aed;">{fmt(tot["mrr"])}</td>'
        f'<td style="padding:9px 14px;text-align:right;font-size:13px;font-weight:800;color:#0891b2;">{fmt(tot["arr"])}</td>'
        f'</tr>'
    )
    return rows


def gerar_tbody_em_analise(em_analise, semaforo_fn):
    rows = ""; tot = {'vtc': 0, 'mrr': 0, 'arr': 0}
    sem = {'🔴': 0, '🟡': 0, '🟢': 0}
    for i, r in enumerate(em_analise):
        dt = parse_dt(r[7])
        emoji, label, cor_txt, cor_bg = semaforo_fn(dt)
        sem[emoji] = sem.get(emoji, 0) + 1
        mk = (MESES[dt.month - 1] + '/26') if dt else '—'
        vtc = pv(r[9]); mrr = pv(r[18]); arr = pv(r[19])
        tot['vtc'] += vtc; tot['mrr'] += mrr; tot['arr'] += arr
        pill = f'<span style="background:{cor_bg};color:{cor_txt};font-size:11px;font-weight:700;padding:3px 10px;border-radius:99px;white-space:nowrap;">{emoji} {label}</span>'
        bg = BG[i % 2]
        rows += (
            f'<tr style="background:{bg};border-bottom:1px solid #f3f4f6;">'
            f'<td style="padding:9px 14px;text-align:center;">{pill}</td>'
            f'<td style="padding:9px 14px;font-family:monospace;text-align:left;font-size:11px;font-weight:600;color:#2563eb;">{str(r[0])}</td>'
            f'<td style="padding:9px 14px;text-align:left;font-size:12px;font-weight:600;color:#111827;">{str(r[2] or "")}</td>'
            f'<td style="padding:9px 14px;text-align:left;font-size:11px;font-weight:600;color:#6b7280;">{str(r[3] or "").split()[0]}</td>'
            f'<td style="padding:9px 14px;text-align:center;font-size:11px;font-weight:600;color:#059669;">{mk}</td>'
            f'<td style="padding:9px 14px;text-align:right;font-size:12px;font-weight:700;color:#111827;">{fmt(vtc)}</td>'
            f'<td style="padding:9px 14px;text-align:right;font-size:12px;font-weight:700;color:#7c3aed;">{brl(mrr)}</td>'
            f'<td style="padding:9px 14px;text-align:right;font-size:12px;font-weight:700;color:#0891b2;">{brl(arr)}</td>'
            f'</tr>'
        )
    n_vm = sem.get('🔴', 0); n_am = sem.get('🟡', 0); n_vd = sem.get('🟢', 0)
    rows += (
        f'<tr style="background:#f8fafc;border-top:1px solid #e5e7eb;">'
        f'<td colspan="8" style="padding:8px 14px;font-size:11px;color:#6b7280;">'
        f'<span style="margin-right:16px;">🔴 <strong>{n_vm}</strong> · &gt;60 dias</span>'
        f'<span style="margin-right:16px;">🟡 <strong>{n_am}</strong> · 31–60 dias</span>'
        f'<span>🟢 <strong>{n_vd}</strong> · ≤30 dias ou futuro</span>'
        f'</td></tr>'
    )
    n = len(em_analise)
    rows += (
        f'<tr style="background:#eff6ff;border-top:2px solid #bfdbfe;">'
        f'<td colspan="5" style="padding:9px 14px;font-size:11px;font-weight:800;text-transform:uppercase;color:#2563eb;">Total · {n} contratos</td>'
        f'<td style="padding:9px 14px;text-align:right;font-size:13px;font-weight:800;color:#111827;">{fmt(tot["vtc"])}</td>'
        f'<td style="padding:9px 14px;text-align:right;font-size:13px;font-weight:800;color:#7c3aed;">{fmt(tot["mrr"])}</td>'
        f'<td style="padding:9px 14px;text-align:right;font-size:13px;font-weight:800;color:#0891b2;">{fmt(tot["arr"])}</td>'
        f'</tr>'
    )
    return rows


def gerar_tbody_req(req_mes):
    rows = ""; tot = {'n': 0, 'vtc': 0, 'mrr': 0, 'arr': 0}
    idx = 0
    for mk in [f'{m}/26' for m in MESES]:
        d = req_mes.get(mk)
        if not d: continue
        bg = BG[idx % 2]; idx += 1
        tot['n'] += d['n']; tot['vtc'] += d['vtc']; tot['mrr'] += d['mrr']; tot['arr'] += d['arr']
        rows += (
            f'<tr style="background:{bg};border-bottom:1px solid #f3f4f6;">'
            f'<td style="padding:10px 16px;font-size:13px;font-weight:700;color:#111827;">{mk}</td>'
            f'<td style="padding:10px 16px;text-align:center;"><span style="background:#f0f2f5;color:#374151;font-size:12px;font-weight:700;padding:3px 12px;border-radius:99px;">{d["n"]}</span></td>'
            f'<td style="padding:10px 16px;text-align:right;font-size:13px;font-weight:700;color:#111827;">{fmt(d["vtc"])}</td>'
            f'<td style="padding:10px 16px;text-align:right;font-size:13px;font-weight:700;color:#7c3aed;">{fmt(d["mrr"]) if d["mrr"] else brl(0)}</td>'
            f'<td style="padding:10px 16px;text-align:right;font-size:13px;font-weight:700;color:#0891b2;">{fmt(d["arr"]) if d["arr"] else brl(0)}</td>'
            f'</tr>'
        )
    n_meses = len(req_mes)
    rows += (
        f'<tr style="background:#f8fafc;border-top:2px solid #e5e7eb;">'
        f'<td style="padding:10px 16px;font-size:11px;font-weight:800;text-transform:uppercase;color:#374151;">Total · {n_meses} meses</td>'
        f'<td style="padding:10px 16px;text-align:center;font-size:13px;font-weight:800;color:#374151;">{tot["n"]}</td>'
        f'<td style="padding:10px 16px;text-align:right;font-size:13px;font-weight:800;color:#111827;">{fmt(tot["vtc"])}</td>'
        f'<td style="padding:10px 16px;text-align:right;font-size:13px;font-weight:800;color:#7c3aed;">{fmt(tot["mrr"])}</td>'
        f'<td style="padding:10px 16px;text-align:right;font-size:13px;font-weight:800;color:#0891b2;">{fmt(tot["arr"])}</td>'
        f'</tr>'
    )
    return rows


def gerar_tbody_cancel_gestao(cancelados_g):
    rows = ""; tot = {'vtc': 0, 'mrr': 0, 'arr': 0}
    for i, r in enumerate(cancelados_g):
        dt_c = parse_dt(r[20])
        mk = (MESES[dt_c.month - 1] + '/26') if dt_c else '—'
        vtc = pv(r[9]); mrr = pv(r[18]); arr = pv(r[19])
        tot['vtc'] += vtc; tot['mrr'] += mrr; tot['arr'] += arr
        bg = "#ffffff" if i % 2 == 0 else "#fef2f2"
        pill = f'<span style="background:#fee2e2;color:#dc2626;font-size:11px;font-weight:700;padding:2px 10px;border-radius:99px;">{mk}</span>'
        rows += (
            f'<tr style="background:{bg};border-bottom:1px solid #fecaca;">'
            f'<td style="padding:9px 14px;font-family:monospace;text-align:left;font-size:11px;font-weight:600;color:#dc2626;">{str(r[0])}</td>'
            f'<td style="padding:9px 14px;text-align:left;font-size:12px;font-weight:600;color:#111827;">{str(r[2] or "")}</td>'
            f'<td style="padding:9px 14px;text-align:left;font-size:11px;font-weight:600;color:#6b7280;">{str(r[3] or "").split()[0]}</td>'
            f'<td style="padding:9px 14px;text-align:center;">{pill}</td>'
            f'<td style="padding:9px 14px;text-align:right;font-size:12px;font-weight:700;color:#111827;">{fmt(vtc)}</td>'
            f'<td style="padding:9px 14px;text-align:right;font-size:12px;font-weight:700;color:#7c3aed;">{brl(mrr)}</td>'
            f'<td style="padding:9px 14px;text-align:right;font-size:12px;font-weight:700;color:#0891b2;">{brl(arr)}</td>'
            f'</tr>'
        )
    n = len(cancelados_g)
    rows += (
        f'<tr style="background:#fef2f2;border-top:2px solid #fecaca;">'
        f'<td colspan="4" style="padding:9px 14px;font-size:11px;font-weight:800;text-transform:uppercase;color:#dc2626;">Total · {n} contratos cancelados</td>'
        f'<td style="padding:9px 14px;text-align:right;font-size:13px;font-weight:800;color:#111827;">{fmt(tot["vtc"])}</td>'
        f'<td style="padding:9px 14px;text-align:right;font-size:13px;font-weight:800;color:#7c3aed;">{fmt(tot["mrr"])}</td>'
        f'<td style="padding:9px 14px;text-align:right;font-size:13px;font-weight:800;color:#0891b2;">{fmt(tot["arr"])}</td>'
        f'</tr>'
    )
    return rows


def gerar_tbody_evolucao(evolucao, mrr_baseline, mrr_ajust):
    rows = ""
    for i, (m, aprov, cancel, red, mrr, delta) in enumerate(evolucao):
        bg = BG[i % 2]
        ap_s = f'<span style="color:#059669;font-weight:600;">+{fmt(aprov)}</span>' if aprov else '<span style="color:#d1d5db;">—</span>'
        rows += (
            f'<tr style="background:{bg};border-bottom:1px solid #f3f4f6;">'
            f'<td style="padding:10px 16px;font-size:13px;font-weight:700;color:#111827;">{m}/26</td>'
            f'<td style="padding:10px 16px;text-align:right;font-size:13px;">{ap_s}</td>'
            f'<td style="padding:10px 16px;text-align:right;font-size:13px;">{dash_cancel(cancel)}</td>'
            f'<td style="padding:10px 16px;text-align:right;font-size:13px;">{dash_red(red)}</td>'
            f'<td style="padding:10px 16px;text-align:right;font-size:13px;font-weight:700;color:#111827;">{fmt(mrr)}</td>'
            f'<td style="padding:10px 16px;text-align:right;font-size:13px;">{dash_delta(delta)}</td>'
            f'</tr>'
        )
    variacao = mrr_ajust - mrr_baseline
    var_s = fmt_n(abs(variacao))
    var_cell = f'<span style="color:#dc2626;font-weight:800;">−R$ {var_s}</span>' if variacao < 0 else f'<span style="color:#059669;font-weight:800;">+R$ {var_s}</span>'
    rows += (
        f'<tr style="background:#eff6ff;border-top:2px solid #bfdbfe;">'
        f'<td style="padding:10px 16px;font-size:11px;font-weight:800;text-transform:uppercase;color:#2563eb;">Baseline</td>'
        f'<td colspan="3" style="padding:10px 16px;text-align:right;font-size:11px;color:#6b7280;">baseline: {fmt(mrr_baseline)}</td>'
        f'<td style="padding:10px 16px;text-align:right;font-size:13px;font-weight:800;color:#2563eb;">{fmt(mrr_ajust)}</td>'
        f'<td style="padding:10px 16px;text-align:right;font-size:12px;">{var_cell}</td>'
        f'</tr>'
    )
    return rows


# ── 5. PREENCHER TEMPLATE ────────────────────────────────────────────────────

def preencher(template, pip, gest, res):
    hoje = date.today().strftime('%d/%m/%Y')
    t = template

    def s(ph, val): return t.replace(f'{{{{{ph}}}}}', str(val))

    # Plural helpers
    def plur(n, s1, s2): return s1 if n == 1 else s2

    n_ent = len(pip['ent_pf']); n_sai = len(pip['sai_pf'])
    s_ent = plur(n_ent, '', 's'); s_sai = plur(n_sai, '', 's')

    # Cor da variação MRR
    cor_var = '#059669' if res['variacao'] >= 0 else '#dc2626'
    sinal_var = '+' if res['variacao'] >= 0 else '−'
    var_abs = fmt_n(abs(res['variacao']))
    sinal_pct = '+' if res['var_pct'] >= 0 else ''

    # Pedro pct
    rd = pip['resp_data']
    n_tot = pip['n_ativos']

    subs = {
        # Pipeline escalares
        'PL_FONTE':        pip['aba_nova'],
        'PL_FONTE_ANT':    pip['aba_ant'] or '',
        'PL_DATA':         hoje,
        'PL_N_ATIVOS':     pip['n_ativos'],
        'PL_N_ENVIADOS':   pip['n_enviados'],
        'PL_N_NEG':        pip['n_neg'],
        'PL_VT_ENVIADOS':  pip['vt_env_fmt'],
        'PL_N_FAT':        pip['n_fat'],
        'PL_VT_FAT':       fmt(pip['vt_fat']),
        'PL_VT_FAT_N':     fmt_n(pip['vt_fat']),
        'PL_MRR_FAT':      fmt(pip['mrr_fat']),
        'PL_ARR_FAT':      fmt(pip['arr_fat']),
        'PL_VT_NEG':       fmt(pip['vt_neg']),
        'PL_N_ENT_PF':     n_ent,
        'PL_N_SAI_PF':     n_sai,
        'PL_S_ENT':        s_ent,
        'PL_S_SAI':        s_sai,
        'PL_VT_ENT_PF':    fmt(sum(pv(r[12]) for r in pip['ent_pf'])),
        # Responsáveis
        'PL_N_PEDRO':      rd['Pedro Schreck']['qtd'],
        'PL_PCT_PEDRO':    rd['Pedro Schreck']['pct'],
        'PL_N_ANDERSON':   rd['Anderson Bamberg']['qtd'],
        'PL_PCT_ANDERSON': rd['Anderson Bamberg']['pct'],
        'PL_N_TALITA':     rd['Talita Rodrigues']['qtd'],
        'PL_PCT_TALITA':   rd['Talita Rodrigues']['pct'],
        'PL_N_INES':       rd['Ines Cristine Nunes Diogo']['qtd'],
        'PL_PCT_INES':     rd['Ines Cristine Nunes Diogo']['pct'],
        # Gestão escalares
        'GT_FONTE':         gest['aba_nova'],
        'GT_DATA':          hoje,
        'GT_N_ATUAL':       gest['n_atual'],
        'GT_N_CANCEL':      gest['n_cancel'],
        'GT_PCT_CANCEL':    gest['pct_cancel'],
        'GT_N_RENOV':       gest['n_renov'],
        'GT_PCT_RENOV':     gest['pct_renov'],
        'GT_VTC_JAN':       fmt(gest['vtc_jan']),
        'GT_VTC_ATUAL':     fmt(gest['vtc_at']),
        'GT_REDUCAO':       fmt(gest['reducao']),
        'GT_REDUCAO_N':     fmt_n(gest['reducao']),
        'GT_PCT_REDUCAO':   round(gest['reducao'] / gest['vtc_jan'] * 100, 1) if gest['vtc_jan'] else 0,
        'GT_TICKET':        fmt(gest['ticket']),
        'GT_MRR_TOTAL':     fmt(gest['total_mrr']),
        'GT_ARR_TOTAL':     fmt(gest['total_arr']),
        'GT_N_MRR':         gest['n_mrr'],
        'GT_N_ARR':         gest['n_arr'],
        'GT_CANCEL_MRR_TOTAL': fmt_n(gest['total_cancel_mrr']),
        'GT_CANCEL_ARR_TOTAL': fmt_n(gest['total_cancel_arr']),
        'GT_RED_MRR_TOTAL':    fmt_n(gest['total_red_mrr']),
        # Resumo escalares
        'RE_MES_AJUST':     res['mes_ref_fmt'],
        'RE_MRR_AJUST':     fmt(res['mrr_ajust']),
        'RE_MRR_VAR':       f'{sinal_var}R$ {var_abs}',
        'RE_MRR_VAR_PCT':   f'{sinal_pct}{res["var_pct"]}',
        'RE_COR_VAR':       cor_var,
        # respData JSON
        'PL_RESP_DATA':     json.dumps(pip['resp_data'], ensure_ascii=False),
    }

    for ph, val in subs.items():
        t = t.replace(f'{{{{{ph}}}}}', str(val))

    # Tbodys
    tbody_map = {
        'PL_TBODY_NEG':       gerar_tbody_neg(pip['neg']),
        'PL_TBODY_FAT_MES':   gerar_tbody_fat_mes(pip['fat_mes'], pip['n_fat'], pip['vt_fat']),
        'PL_TBODY_DETALHE':   gerar_tbody_detalhe(pip['fechados'], pip['n_fat'], pip['vt_fat'], pip['mrr_fat'], pip['arr_fat']),
        'PL_TBODY_ENT_PF':    gerar_tbody_ent_pf(pip['ent_pf']),
        'PL_TBODY_SAI_PF':    gerar_tbody_sai_pf(pip['sai_pf'], pip['aba_ant'], pip['aba_nova']),
        'GT_TBODY_MRR':       gerar_tbody_mrr(gest['mrr_mes'], gest['total_mrr'], gest['total_cancel_mrr'], gest['total_red_mrr'], gest['n_mrr']),
        'GT_TBODY_ARR':       gerar_tbody_arr(gest['arr_mes'], gest['total_arr'], gest['total_cancel_arr'], gest['total_red_arr'], gest['n_arr']),
        'GT_TBODY_RENOV':     gerar_tbody_gestao_tabela(gest['renovados'], '#059669', ('#f0fdf4','#bbf7d0','#059669'), 'Total'),
        'GT_TBODY_PF':        gerar_tbody_gestao_tabela(gest['pf_g'], '#7c3aed', ('#f5f3ff','#e9d5ff','#7c3aed'), 'Total'),
        'GT_TBODY_ANALISE':   gerar_tbody_em_analise(gest['em_analise'], gest['semaforo_fn']),
        'GT_TBODY_REQ':       gerar_tbody_req(gest['req_mes']),
        'GT_TBODY_CANCEL':    gerar_tbody_cancel_gestao(gest['cancelados_g']),
        'RE_TBODY_EVOLUCAO':  gerar_tbody_evolucao(res['evolucao'], res['mrr_baseline'], res['mrr_ajust']),
        'RE_TBODY_FAT_MES':   gerar_tbody_fat_mes(pip['fat_mes'], pip['n_fat'], pip['vt_fat']),
        'RE_TBODY_DETALHE':   gerar_tbody_detalhe(pip['fechados'], pip['n_fat'], pip['vt_fat'], pip['mrr_fat'], pip['arr_fat']),
    }

    for ph, html in tbody_map.items():
        t = t.replace(f'<tbody>{{{{{ph}}}}}  </tbody>', f'<tbody>{html}</tbody>')
        t = t.replace(f'<tbody>{{{{{ph}}}}}</tbody>', f'<tbody>{html}</tbody>')

    return t


# ── 6. VALIDAÇÃO ─────────────────────────────────────────────────────────────

def validar(worker, pip, gest):
    issues = []
    checks = [
        (fmt(pip['vt_fat']), f"VT Fechados {fmt(pip['vt_fat'])}"),
        (fmt(pip['mrr_fat']), f"MRR Fechados {fmt(pip['mrr_fat'])}"),
        (fmt(gest['total_mrr']), f"MRR Total gestão {fmt(gest['total_mrr'])}"),
        (str(pip['n_fat']), f"N fechados {pip['n_fat']}"),
        (str(gest['n_cancel']), f"N cancelados {gest['n_cancel']}"),
        (pip['aba_nova'], f"Fonte {pip['aba_nova']}"),
    ]
    for val, desc in checks:
        if val not in worker:
            issues.append(f"  ⚠️  {desc} — não encontrado no worker gerado")

    remaining = re.findall(r'\{\{[A-Z_]+\}\}', worker)
    if remaining:
        issues.append(f"  ⚠️  Placeholders não preenchidos: {set(remaining)}")

    div_open  = len(re.findall(r'<div[\s>]', worker))
    div_close = worker.count('</div>')
    if div_open - div_close not in (0, -1):
        issues.append(f"  ⚠️  Div desbalanceado: {div_open} open, {div_close} close")

    return issues


# ── 7. MAIN ───────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description='Atualiza o dashboard BI Qualitor')
    ap.add_argument('--projetos',  required=True, help='Relatório_de_Projetos.xlsx')
    ap.add_argument('--contratos', required=True, help='Renovação_de_Contratos.xlsx')
    ap.add_argument('--reducao',   default=None,  help='Redução_Contratual.xlsx (opcional)')
    ap.add_argument('--template',  required=True, help='template.js')
    ap.add_argument('--saida',     default='worker.js', help='worker.js de saída')
    args = ap.parse_args()

    print("\n🔄 Dashboard BI Qualitor — Atualização Automática")
    print("=" * 50)

    print("\n📊 Calculando Pipeline...")
    pip = calcular_pipeline(args.projetos)

    print("\n📋 Calculando Gestão de Contratos...")
    gest = calcular_gestao(args.contratos, args.reducao)

    print("\n📈 Calculando Resumo Executivo...")
    res = calcular_resumo(pip, gest)

    print("\n📝 Sumário:")
    print(f"  Pipeline  → {pip['aba_ant']} → {pip['aba_nova']} | Ativos={pip['n_ativos']} | Fechados={pip['n_fat']} | VT={fmt(pip['vt_fat'])}")
    print(f"  Gestão    → {gest['aba_nova']} | Carteira={gest['n_atual']} | Cancelados={gest['n_cancel']} | MRR={fmt(gest['total_mrr'])}")
    print(f"  Resumo    → Baseline={fmt(res['mrr_baseline'])} | Ajustado={fmt(res['mrr_ajust'])} | Var={res['var_pct']:+.2f}%")

    print("\n🔧 Preenchendo template...")
    with open(args.template, 'r', encoding='utf-8') as f:
        template = f.read()

    worker = preencher(template, pip, gest, res)

    print("\n✅ Validando...")
    issues = validar(worker, pip, gest)
    if issues:
        print("  Problemas encontrados:")
        for i in issues: print(i)
    else:
        print("  Tudo OK — nenhum problema encontrado")

    with open(args.saida, 'w', encoding='utf-8') as f:
        f.write(worker)

    print(f"\n✅ worker.js gerado: {args.saida} ({len(worker):,} bytes)")
    print("   → Faça push no GitHub para deploy automático\n")


if __name__ == '__main__':
    main()
